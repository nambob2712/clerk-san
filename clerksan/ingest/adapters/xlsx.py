"""Bounded spreadsheet normalization for SQL-oriented workbook ingestion."""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import re
import warnings
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, UnidentifiedImageError

from clerksan.ingest.filetype import FileType
from clerksan.ingest.limits import (
    IngestLimits,
    ResourceLimitExceeded,
    check_image_pixels,
    safe_zip_members,
)
from clerksan.ingest.normalized import (
    DocMetadata,
    ExtractedImage,
    ExtractedTable,
    NormalizedDocument,
)
from clerksan.ingest.parser_runner import AdapterContext, ReadOnlySource

from .source_io import read_bounded_source

_MAX_WORKSHEET_CELLS = 1_000_000
_MAX_SHEET_DESCRIPTION_CHARS = 1_000
_MIN_EMBEDDED_IMAGE_DIMENSION = 50
_HEADER_KEYWORDS = frozenset(
    {
        "date",
        "amount",
        "total",
        "vendor",
        "counterparty",
        "description",
        "日付",
        "金額",
        "合計",
        "取引先",
        "摘要",
        "備考",
        "科目",
    }
)


def detect_header_row(rows: list[list[str]]) -> int:
    """Return the most table-like header row, falling back to the first nonblank row."""

    best_index = 0
    best_score = float("-inf")
    for index, row in enumerate(rows):
        cells = [cell.strip() for cell in row if cell.strip()]
        if not cells:
            continue
        distinct = {cell.casefold() for cell in cells}
        text_cells = sum(not _looks_like_data(cell) for cell in cells)
        keyword_hits = sum(_is_header_keyword(cell) for cell in cells)
        has_following_data = any(
            sum(bool(cell.strip()) for cell in candidate) >= min(2, len(cells))
            for candidate in rows[index + 1 :]
        )
        score = (
            len(cells)
            + (2 * len(distinct))
            + text_cells
            + (4 * keyword_hits)
            + (2 if has_following_data else 0)
            - (index / 1000)
        )
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def split_subtables(rows: list[list[str]]) -> list[list[list[str]]]:
    """Split an expanded worksheet grid at completely blank row and column boundaries."""

    if not rows:
        return []
    width = max((len(row) for row in rows), default=0)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    subtables: list[list[list[str]]] = []

    for row_start, row_end in _nonblank_ranges(normalized, axis="row"):
        block = normalized[row_start:row_end]
        for column_start, column_end in _nonblank_ranges(block, axis="column"):
            table = [
                row[column_start:column_end]
                for row in block
                if any(cell.strip() for cell in row[column_start:column_end])
            ]
            if table:
                subtables.append(table)
    return subtables


class XlsxAdapter:
    """Extract workbook tables without making spreadsheet rows semantic chunks."""

    supported_types: tuple[FileType, ...] = (FileType.XLSX,)

    def __init__(self, *, limits: IngestLimits | None = None) -> None:
        self.limits = limits or IngestLimits()

    def normalize(self, source: ReadOnlySource, context: AdapterContext) -> NormalizedDocument:
        detected_type = _context_file_type(context)
        raw = read_bounded_source(source, self.limits)
        meta = DocMetadata(
            filename=source.filename,
            detected_type=detected_type,
            sha256=source.source_sha256,
            family="tabular",
            canonical_mime=_context_text(context, "canonical_mime"),
        )
        return self._normalize_bytes(raw, meta)

    async def adapt(self, raw: bytes, meta: DocMetadata) -> NormalizedDocument:
        """Validate the OOXML archive before loading cells and merged ranges."""

        return self._normalize_bytes(raw, meta)

    def _normalize_bytes(self, raw: bytes, meta: DocMetadata) -> NormalizedDocument:
        """Normalize workbook bytes without formula evaluation or write-mode loading."""

        if meta.detected_type is not FileType.XLSX:
            raise ValueError(f"XLSX adapter cannot handle {meta.detected_type.value!r}")
        workbook = _load_checked_workbook(raw, self.limits)
        try:
            tables, descriptions = _extract_workbook_tables(
                workbook, meta, _merged_ranges_by_sheet(raw, self.limits)
            )
        finally:
            workbook.close()

        images = extract_embedded_images(raw, limits=self.limits)
        extra = dict(meta.extra)
        extra.update(
            {
                "document_format": "xlsx",
                "sheet_descriptions": descriptions,
                "spreadsheet_row_embedding": "disabled",
                "embedded_image_count": len(images),
                "table_count": len(tables),
            }
        )
        body = "\n\n".join(descriptions)
        return NormalizedDocument(
            markdown_body=body,
            metadata=meta.model_copy(update={"extra": extra}),
            images=images,
            tables=tables,
            embeddable=False,
        )


def sheet_descriptions(
    raw: bytes, meta: DocMetadata, *, limits: IngestLimits | None = None
) -> list[str]:
    """Return one bounded, row-free discovery description per worksheet."""

    if meta.detected_type is not FileType.XLSX:
        raise ValueError(f"XLSX descriptions cannot handle {meta.detected_type.value!r}")
    workbook = _load_checked_workbook(raw, limits or IngestLimits())
    try:
        _, descriptions = _extract_workbook_tables(
            workbook, meta, _merged_ranges_by_sheet(raw, limits or IngestLimits())
        )
        return descriptions
    finally:
        workbook.close()


def extract_embedded_images(
    xlsx_bytes: bytes, *, limits: IngestLimits | None = None
) -> list[ExtractedImage]:
    """Return distinct non-trivial ``xl/media`` images for a downstream OCR job."""

    active_limits = limits or IngestLimits()
    images: list[ExtractedImage] = []
    seen: set[str] = set()
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as archive:
            safe_zip_members(archive, active_limits)
            for member in archive.infolist():
                if member.is_dir() or not member.filename.startswith("xl/media/"):
                    continue
                image_bytes = archive.read(member)
                dimensions = _embedded_image_dimensions(image_bytes, active_limits)
                if dimensions is None:
                    continue
                width, height = dimensions
                if width < _MIN_EMBEDDED_IMAGE_DIMENSION or height < _MIN_EMBEDDED_IMAGE_DIMENSION:
                    continue
                digest = hashlib.sha256(image_bytes).hexdigest()
                if digest in seen:
                    continue
                seen.add(digest)
                images.append(
                    ExtractedImage(
                        sha256=digest,
                        content_path=f"embedded/sha256/{digest}",
                        width=width,
                        height=height,
                        source_location=member.filename,
                    )
                )
    except zipfile.BadZipFile as error:
        raise ValueError("invalid XLSX archive") from error
    return images


def _load_checked_workbook(raw: bytes, limits: IngestLimits):
    _validate_xlsx_archive(raw, limits)
    try:
        return load_workbook(
            io.BytesIO(raw),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as error:
        raise ValueError("invalid or unreadable XLSX upload") from error


def _validate_xlsx_archive(raw: bytes, limits: IngestLimits) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            safe_zip_members(archive, limits)
    except zipfile.BadZipFile as error:
        raise ValueError("invalid XLSX archive") from error


def _merged_ranges_by_sheet(raw: bytes, limits: IngestLimits) -> dict[str, tuple[str, ...]]:
    """Read only merge declarations omitted by openpyxl's streaming worksheet API."""

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            safe_zip_members(archive, limits)
            workbook_xml = _safe_xml_member(archive, "xl/workbook.xml")
            relationships_xml = _safe_xml_member(archive, "xl/_rels/workbook.xml.rels")
            workbook_root = ET.fromstring(workbook_xml)
            relationships_root = ET.fromstring(relationships_xml)
            relationship_targets = {
                element.attrib.get("Id", ""): element.attrib.get("Target", "")
                for element in relationships_root
                if element.tag.rsplit("}", 1)[-1] == "Relationship"
            }
            result: dict[str, tuple[str, ...]] = {}
            for sheet in workbook_root.iter():
                if sheet.tag.rsplit("}", 1)[-1] != "sheet":
                    continue
                title = sheet.attrib.get("name", "")
                relationship_id = next(
                    (
                        value
                        for key, value in sheet.attrib.items()
                        if key.rsplit("}", 1)[-1] == "id"
                    ),
                    "",
                )
                target = relationship_targets.get(relationship_id, "")
                if not title or not target:
                    raise ValueError("XLSX worksheet relationship is incomplete")
                member = target.lstrip("/")
                if not member.startswith("xl/"):
                    member = f"xl/{member}"
                sheet_root = ET.fromstring(_safe_xml_member(archive, member))
                ranges = tuple(
                    element.attrib["ref"]
                    for element in sheet_root.iter()
                    if element.tag.rsplit("}", 1)[-1] == "mergeCell" and "ref" in element.attrib
                )
                result[title] = ranges
            return result
    except (KeyError, ET.ParseError, OSError, zipfile.BadZipFile) as error:
        raise ValueError("invalid XLSX merge metadata") from error


def _safe_xml_member(archive: zipfile.ZipFile, name: str) -> bytes:
    info = archive.getinfo(name)
    if info.file_size > 8 * 1024 * 1024:
        raise ValueError("XLSX metadata XML exceeds its bound")
    data = archive.read(info)
    uppercase = data[:4096].upper()
    if b"<!DOCTYPE" in uppercase or b"<!ENTITY" in uppercase:
        raise ValueError("XLSX metadata XML contains a forbidden declaration")
    return data


def _extract_workbook_tables(
    workbook,
    meta: DocMetadata,
    merged_ranges_by_sheet: dict[str, tuple[str, ...]],
) -> tuple[list[ExtractedTable], list[str]]:
    tables: list[ExtractedTable] = []
    descriptions: list[str] = []
    for sheet in workbook.worksheets:
        grid = _worksheet_grid(sheet, merged_ranges_by_sheet.get(sheet.title, ()))
        sheet_tables = split_subtables(grid)
        headers: list[str] = []
        table_number = 0
        for matrix in sheet_tables:
            header_index = detect_header_row(matrix)
            header = _unique_headers(matrix[header_index])
            rows = [row for row in matrix[header_index + 1 :] if any(cell.strip() for cell in row)]
            if not rows:
                continue
            table_number += 1
            tables.append(
                ExtractedTable(
                    header=header,
                    rows=rows,
                    source_location=f"sheet:{sheet.title}:table:{table_number}",
                )
            )
            headers.extend(header)
        descriptions.append(_sheet_description(meta.filename, sheet, headers))
    return tables, descriptions


def _worksheet_grid(sheet: Worksheet, merged_ranges: tuple[str, ...] = ()) -> list[list[str]]:
    cell_count = sheet.max_row * sheet.max_column
    if cell_count > _MAX_WORKSHEET_CELLS:
        raise ResourceLimitExceeded("max_worksheet_cells", _MAX_WORKSHEET_CELLS, cell_count)
    grid = [
        [_cell_text(cell.value) for cell in row]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
    ]
    for merged_range in merged_ranges:
        min_column, min_row, max_column, max_row = range_boundaries(merged_range)
        if (
            min_row < 1
            or min_column < 1
            or max_row > len(grid)
            or (grid and max_column > len(grid[0]))
        ):
            raise ValueError("XLSX merge range exceeds worksheet dimensions")
        value = grid[min_row - 1][min_column - 1]
        for row_index in range(min_row - 1, max_row):
            for column_index in range(min_column - 1, max_column):
                grid[row_index][column_index] = value
    return grid


def _nonblank_ranges(rows: list[list[str]], *, axis: str) -> list[tuple[int, int]]:
    if axis == "row":
        values = [any(cell.strip() for cell in row) for row in rows]
    elif axis == "column":
        values = [any(row[index].strip() for row in rows) for index in range(len(rows[0]))]
    else:
        raise ValueError("axis must be row or column")

    ranges: list[tuple[int, int]] = []
    start: int | None = None
    for index, is_nonblank in enumerate(values):
        if is_nonblank and start is None:
            start = index
        elif not is_nonblank and start is not None:
            ranges.append((start, index))
            start = None
    if start is not None:
        ranges.append((start, len(values)))
    return ranges


def _unique_headers(values: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = value.strip() or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        suffix = "" if counts[base] == 1 else f"_{counts[base]}"
        headers.append(f"{base}{suffix}")
    return headers


def _sheet_description(filename: str, sheet: Worksheet, headers: Iterable[str]) -> str:
    unique_headers = list(dict.fromkeys(header for header in headers if header))
    columns = ", ".join(unique_headers[:25]) or "no detected columns"
    dates = [
        value
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if isinstance(value, (dt.date, dt.datetime))
    ]
    description = f"Workbook {filename}; sheet {sheet.title}. Columns: {columns}."
    if dates:
        normalized_dates = [
            value.date() if isinstance(value, dt.datetime) else value for value in dates
        ]
        date_start = min(normalized_dates).isoformat()
        date_end = max(normalized_dates).isoformat()
        description += f" Date range: {date_start} to {date_end}."
    return description[:_MAX_SHEET_DESCRIPTION_CHARS]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        if value.time() == dt.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def _looks_like_data(value: str) -> bool:
    compact = value.replace(",", "").replace("円", "").replace("%", "")
    return bool(
        re.fullmatch(r"[+-]?\d+(?:\.\d+)?", compact)
        or re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", value)
    )


def _is_header_keyword(value: str) -> bool:
    normalized = value.casefold()
    return any(keyword.casefold() in normalized for keyword in _HEADER_KEYWORDS)


def _embedded_image_dimensions(data: bytes, limits: IngestLimits) -> tuple[int, int] | None:
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            with Image.open(io.BytesIO(data)) as image:
                width, height = image.size
                check_image_pixels(width, height, limits)
                image.verify()
                return width, height
    except (UnidentifiedImageError, OSError):
        return None
    except (Image.DecompressionBombError, Image.DecompressionBombWarning) as error:
        raise ValueError("embedded image exceeds Pillow safety limits") from error


def _context_file_type(context: AdapterContext) -> FileType:
    try:
        return FileType(context.metadata.get("detected_type", ""))
    except ValueError as error:
        raise ValueError("XLSX adapter context requires a detected type") from error


def _context_text(context: AdapterContext, key: str) -> str | None:
    value = context.metadata.get(key)
    return value if isinstance(value, str) else None
