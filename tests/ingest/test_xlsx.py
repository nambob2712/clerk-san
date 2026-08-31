from __future__ import annotations

import datetime as dt
import hashlib
import io
import zipfile

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as SpreadsheetImage
from PIL import Image

from clerksan.ingest.adapters.xlsx import (
    XlsxAdapter,
    detect_header_row,
    sheet_descriptions,
    split_subtables,
)
from clerksan.ingest.filetype import FileType
from clerksan.ingest.limits import UnsafeArchiveMemberError
from clerksan.ingest.normalized import DocMetadata


def _png() -> io.BytesIO:
    image = Image.new("RGB", (60, 60), color="teal")
    output = io.BytesIO()
    image.save(output, format="PNG")
    image.close()
    output.seek(0)
    return output


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "経費"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "2026年 経費レポート"
    sheet.append([])
    sheet.append(["日付", "金額", "摘要", None, "商品", "数量"])
    sheet.append([dt.date(2026, 7, 1), 1200, "交通費", None, "ノート", 2])
    sheet.append([dt.date(2026, 7, 2), 500, None, None, "ペン", 3])
    sheet.merge_cells("C4:C5")
    sheet.add_image(SpreadsheetImage(_png()), "H2")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _metadata(raw: bytes) -> DocMetadata:
    return DocMetadata(
        filename="expenses.xlsx",
        detected_type=FileType.XLSX,
        sha256=hashlib.sha256(raw).hexdigest(),
        extra={"content_path": "documents/expenses.xlsx"},
    )


def test_header_and_subtable_helpers_handle_title_rows_and_blank_separators() -> None:
    rows = [
        ["経費レポート", "経費レポート", "経費レポート", "", "", ""],
        ["", "", "", "", "", ""],
        ["日付", "金額", "摘要", "", "商品", "数量"],
        ["2026-07-01", "1200", "交通費", "", "ノート", "2"],
    ]

    assert detect_header_row(rows) == 2
    subtables = split_subtables(rows[2:])
    assert subtables == [
        [["日付", "金額", "摘要"], ["2026-07-01", "1200", "交通費"]],
        [["商品", "数量"], ["ノート", "2"]],
    ]


async def test_xlsx_adapter_fills_merges_extracts_subtables_and_disables_embedding() -> None:
    raw = _workbook_bytes()
    result = await XlsxAdapter().adapt(raw, _metadata(raw))

    assert result.embeddable is False
    assert len(result.tables) == 2
    assert result.tables[0].header == ["日付", "金額", "摘要"]
    assert result.tables[0].rows == [
        ["2026-07-01", "1200", "交通費"],
        ["2026-07-02", "500", "交通費"],
    ]
    assert result.tables[1].header == ["商品", "数量"]
    assert result.tables[1].rows == [["ノート", "2"], ["ペン", "3"]]
    assert "1200" not in result.markdown_body
    assert result.metadata.extra["spreadsheet_row_embedding"] == "disabled"
    assert len(result.images) == 1
    assert result.images[0].source_location == "xl/media/image1.png"


def test_xlsx_sheet_descriptions_are_one_per_sheet_and_exclude_rows() -> None:
    raw = _workbook_bytes()
    descriptions = sheet_descriptions(raw, _metadata(raw))

    assert len(descriptions) == 1
    assert "sheet 経費" in descriptions[0]
    assert "日付" in descriptions[0]
    assert "2026-07-01 to 2026-07-02" in descriptions[0]
    assert "1200" not in descriptions[0]


async def test_xlsx_rejects_unsafe_archive_member_before_openpyxl() -> None:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr("../outside.xml", "not a workbook")
    raw = output.getvalue()

    with pytest.raises(UnsafeArchiveMemberError):
        await XlsxAdapter().adapt(raw, _metadata(raw))
